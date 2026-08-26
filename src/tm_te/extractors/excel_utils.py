"""Low-level, generic helpers for reading the loosely-structured xlsx inputs.

Every real workbook in this project mixes cover pages, multi-row titles and
data tables that don't always start in row/column 1, so we read with
``openpyxl`` directly instead of assuming pandas' "row 0 is the header"
default.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Iterable, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

TRUTHY_STRINGS = {"true", "yes", "y", "1", "o", "checked", "x"}
VALID_MARKER_STRINGS = {"o"}
INVALID_MARKER_STRINGS = {"x"}


def load_workbook(path: str):
    """Load a workbook read-only, with computed values instead of formulas."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Input workbook not found: {path}")
    return openpyxl.load_workbook(p, data_only=True, read_only=False)


def get_sheet(wb, sheet_name: str) -> Worksheet:
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )
    return wb[sheet_name]


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_lower(value: Any) -> str:
    return _norm(value).lower()


def is_valid_marker(value: Any) -> bool:
    """True when a Comm-Matrix / App-Parameter / IO-Signal cell means 'O' (valid)."""
    return norm_lower(value) in VALID_MARKER_STRINGS


def is_truthy(value: Any) -> bool:
    """True for checkbox-like cells (booleans, 'TRUE', 'Y', 1, ...)."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    return norm_lower(value) in TRUTHY_STRINGS


def find_header_row(
    ws: Worksheet, target: str, search_rows: int = 15, max_col: Optional[int] = None
) -> Optional[int]:
    """Scan the first ``search_rows`` rows for a cell that equals ``target``
    (case-insensitive, whitespace-trimmed). Returns the 1-indexed row number.

    This is how sheets like "Master Comm Matrix (CAN)" locate the column for
    a given requirement-sheet id ("019"), and how the signal-feature sheets
    locate the column for a given feature id.
    """
    target_norm = norm_lower(target)
    if not target_norm:
        return None
    max_col = max_col or ws.max_column
    max_row = min(search_rows, ws.max_row)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if norm_lower(ws.cell(row=row, column=col).value) == target_norm:
                return row
    return None


def find_column_in_row(ws: Worksheet, header_row: int, target: str) -> Optional[int]:
    """Find the 1-indexed column whose cell in ``header_row`` equals ``target``."""
    target_norm = norm_lower(target)
    for col in range(1, ws.max_column + 1):
        if norm_lower(ws.cell(row=header_row, column=col).value) == target_norm:
            return col
    return None


def find_column_containing(ws: Worksheet, header_row: int, needle: str) -> Optional[int]:
    """Find the first column in ``header_row`` whose header *contains* ``needle``."""
    needle_norm = norm_lower(needle)
    for col in range(1, ws.max_column + 1):
        if needle_norm in norm_lower(ws.cell(row=header_row, column=col).value):
            return col
    return None


def header_names(ws: Worksheet, header_row: int, max_col: Optional[int] = None) -> list[str]:
    max_col = max_col or ws.max_column
    return [_norm(ws.cell(row=header_row, column=c).value) for c in range(1, max_col + 1)]


def sheet_to_records(
    ws: Worksheet,
    header_row: int = 1,
    max_col: Optional[int] = None,
    start_row: Optional[int] = None,
) -> list[dict[str, Any]]:
    """Read every row below ``header_row`` into a list of dicts keyed by the
    header row's cell values. Fully-blank rows are skipped."""
    headers = header_names(ws, header_row, max_col)
    max_col = max_col or ws.max_column
    records = []
    for row in range(start_row or header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
        if all(v is None or _norm(v) == "" for v in values):
            continue
        record = {
            headers[i] if headers[i] else f"col_{i + 1}": values[i]
            for i in range(len(values))
        }
        records.append(record)
    return records


def filter_records_by_marker(
    ws: Worksheet,
    header_row: int,
    marker_col: int,
    is_valid: Callable[[Any], bool],
    keep_cols: Optional[Iterable[int]] = None,
) -> list[dict[str, Any]]:
    """Filter data rows where ``marker_col`` passes ``is_valid``, keeping only
    ``keep_cols`` (1-indexed, defaults to all columns) plus the marker value
    itself under the marker's own header name."""
    headers = header_names(ws, header_row)
    keep_cols = list(keep_cols) if keep_cols is not None else list(range(1, ws.max_column + 1))
    if marker_col not in keep_cols:
        keep_cols = keep_cols + [marker_col]

    results = []
    for row in range(header_row + 1, ws.max_row + 1):
        marker_value = ws.cell(row=row, column=marker_col).value
        if not is_valid(marker_value):
            continue
        record: dict[str, Any] = {}
        for c in keep_cols:
            key = headers[c - 1] if c - 1 < len(headers) and headers[c - 1] else f"col_{c}"
            record[key] = ws.cell(row=row, column=c).value
        if all(v is None or _norm(v) == "" for v in record.values()):
            continue
        results.append(record)
    return results


def find_row_by_value(
    ws: Worksheet, header_row: int, column: int, value: str
) -> Optional[dict[str, Any]]:
    """Return the first data row (as a dict) below ``header_row`` whose cell in
    ``column`` matches ``value`` (case-insensitive, trimmed)."""
    target = norm_lower(value)
    headers = header_names(ws, header_row)
    for row in range(header_row + 1, ws.max_row + 1):
        if norm_lower(ws.cell(row=row, column=column).value) == target:
            record = {}
            for c in range(1, ws.max_column + 1):
                key = headers[c - 1] if c - 1 < len(headers) and headers[c - 1] else f"col_{c}"
                record[key] = ws.cell(row=row, column=c).value
            return record
    return None


def row_blocks(ws: Worksheet, first_col: int = 1, last_col: Optional[int] = None):
    """Yield lists of raw row-value-tuples, grouped by blank-row separators.

    Used for sheets like "Compound Commands (Set)" where multiple small
    tables are stacked in one sheet, each preceded/followed by a blank row.
    """
    last_col = last_col or ws.max_column
    current: list[list[Any]] = []
    for row in range(1, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(first_col, last_col + 1)]
        blank = all(v is None or _norm(v) == "" for v in values)
        if blank:
            if current:
                yield current
                current = []
            continue
        current.append(values)
    if current:
        yield current
