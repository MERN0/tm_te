"""Generates small synthetic xlsx workbooks that mimic the real input
files' documented structure, so Phase 1 can be exercised end-to-end without
the real (large, confidential) source spreadsheets.

Run: python tests/fixtures/build_fixtures.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).parent


def _write_rows(ws, rows, start_row=1):
    for r, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)


def build_system_requirements():
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover Page")
    cover["A1"] = "System Requirements"

    index = wb.create_sheet("Index")
    _write_rows(index, [
        ["Req Sheet", "Feature Name"],
        ["019", "Adaptive Cruise Control"],
        ["010", "Lane Keep Assist"],
        ["001", "Parking Brake"],
    ])

    s019 = wb.create_sheet("019")
    _write_rows(s019, [
        ["Req ID", "Requirement Type", "Requirement Description"],
        ["019-001", "Functional Requirement", "System shall maintain set speed"],
        ["019-002", "Non-Functional Requirement", "System shall respond within 50ms"],
        ["019-003", "Functional Requirement", "System shall brake on obstacle detection"],
    ])

    s010 = wb.create_sheet("010")
    _write_rows(s010, [
        ["Req ID", "Requirement Type", "Requirement Description"],
        ["010-001", "Functional Requirement", "System shall detect lane markings"],
    ])

    comm_matrix = wb.create_sheet("Master Comm Matrix (CAN)")
    _write_rows(comm_matrix, [
        ["Signal Name", "Description", "019", "010", "001"],
        ["ACC_SetSpeed", "Set speed command", "O", "X", "X"],
        ["ACC_Enable", "Enable ACC", "O", "X", "X"],
        ["LKA_LaneOffset", "Lane offset", "X", "O", "X"],
        ["PB_State", "Parking brake state", "X", "X", "O"],
    ])

    app_param = wb.create_sheet("Master List - App Parameter")
    _write_rows(app_param, [
        ["Parameter Name", "Description", "019", "010", "001"],
        ["ACC_MaxAccel", "Maximum acceleration", "O", "X", "X"],
        ["LKA_Sensitivity", "Lane sensitivity", "X", "O", "X"],
    ])

    io_signals = wb.create_sheet("Master Input Output Signals")
    _write_rows(io_signals, [
        ["Signal Name", "Direction", "019", "010", "001"],
        ["ACC_RadarDistance", "Input", "O", "X", "X"],
        ["ACC_ThrottleCmd", "Output", "O", "X", "X"],
        ["PB_ApplyCmd", "Output", "X", "X", "O"],
    ])

    abbrev = wb.create_sheet("Master List Abbreviations")
    _write_rows(abbrev, [["Abbreviation", "Meaning"], ["ACC", "Adaptive Cruise Control"]])

    wb.save(FIXTURES_DIR / "System Requirements.xlsx")


def build_command_list():
    wb = Workbook()
    wb.remove(wb.active)

    mapping = wb.create_sheet("Feature_ID_Mapping")
    _write_rows(mapping, [
        ["Feature Name", "ID"],
        ["Adaptive Cruise Control", 3],
        ["Lane Keep Assist", 4],
        ["Parking Brake", 5],
    ])

    recorder = wb.create_sheet("Recorder_Signal_Feature")
    _write_rows(recorder, [
        ["Signal Name", "Command Name", "Description", "Unit", "3", "4", "5"],
        ["ACC_SetSpeed_Rec", "SetCruiseSpeed", "Records set speed", "km/h", True, False, False],
        ["ACC_Enable_Rec", "EnableACC", "Records enable state", "bool", True, False, False],
        ["LKA_Offset_Rec", "SetLaneOffset", "Records offset", "m", False, True, False],
    ])

    recorder_diag = wb.create_sheet("Recorder_Signal_Diag")
    _write_rows(recorder_diag, [["Signal Name", "DTC"], ["ACC_Fault", "P0500"]])

    sdo_feature = wb.create_sheet("SDO_Signal_Feature")
    _write_rows(sdo_feature, [
        ["Signal Name", "Command Name", "Description", "Unit", "3", "4", "5"],
        ["ACC_Brake_SDO", "ApplyBrake", "Braking SDO", "bool", True, False, False],
        ["PB_State_SDO", "SetParkingBrake", "Parking brake SDO", "bool", False, False, True],
    ])

    sdo_diag = wb.create_sheet("SDO_Signal_Diag")
    _write_rows(sdo_diag, [["Signal Name", "DTC"], ["PB_Fault", "P0600"]])

    command_list = wb.create_sheet("Command List")
    _write_rows(command_list, [
        ["Command Name", "Command Type", "Description", "Parameters"],
        ["SetCruiseSpeed", "Set", "Sets the ACC cruise speed", "speed_kph"],
        ["EnableACC", "Set", "Enables ACC", "enable_flag"],
        ["ApplyBrake", "Set", "Applies braking force", "brake_pct"],
        ["SetLaneOffset", "Set", "Sets lane offset target", "offset_m"],
        ["SetParkingBrake", "Set", "Applies parking brake", "state"],
    ])

    wb.save(FIXTURES_DIR / "TE_TMHC_Command_List.xlsx")


def build_configuration_file():
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover Page")
    cover["A1"] = "Configuration File"

    tolerances = wb.create_sheet("Tolerances")
    _write_rows(tolerances, [
        ["Parameter", "Tolerance", "Unit"],
        ["Speed", 0.5, "km/h"],
        ["Brake Pressure", 2.0, "bar"],
    ])

    model_input = wb.create_sheet("Model_Input_Mapping")
    _write_rows(model_input, [
        ["Command Name", "Model Input Signal", "Scale"],
        ["SetCruiseSpeed", "Model.ACC.SetSpeed", 1.0],
        ["EnableACC", "Model.ACC.Enable", 1.0],
        ["ApplyBrake", "Model.Brake.Cmd", 1.0],
    ])

    wb.save(FIXTURES_DIR / "TE_TMHC_Configuration_File.xlsx")


def build_compound_commands():
    wb = Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover Page")
    cover["A1"] = "Compound Commands"

    set_sheet = wb.create_sheet("Compound Commands (Set)")
    _write_rows(set_sheet, [
        ["SetCruiseAndEnable"],
        ["Step", "Command", "Value"],
        [1, "SetCruiseSpeed", "100"],
        [2, "EnableACC", "1"],
    ], start_row=1)
    _write_rows(set_sheet, [
        ["ApplyFullBrake"],
        ["Step", "Command", "Value"],
        [1, "ApplyBrake", "100"],
    ], start_row=6)

    verify_sheet = wb.create_sheet("Compound Commands (Verify)")
    _write_rows(verify_sheet, [
        ["VerifyCruiseActive"],
        ["Step", "Signal", "Expected"],
        [1, "ACC_Enable_Rec", "1"],
        [2, "ACC_SetSpeed_Rec", "100"],
    ])

    wb.save(FIXTURES_DIR / "TE_TMHC_Compound_Commands.xlsx")


def build_library_description():
    wb = Workbook()
    wb.remove(wb.active)

    lib_list = wb.create_sheet("Library List")
    lib_list["C2"] = "Section: Core Libraries"
    lib_list["C4"] = "CANLibrary"
    lib_list["D4"] = "CAN bus signal read/write helpers"
    lib_list["E4"] = "CANLibrary.Send(signal, value)"
    lib_list["C6"] = "ModelLibrary"
    lib_list["D6"] = "Simulink model interaction helpers"
    lib_list["E6"] = "ModelLibrary.SetInput(name, value)"

    custom_kw = wb.create_sheet("Custom Keyword&Library Details")
    custom_kw["C2"] = "Section: Custom Keywords"
    custom_kw["C4"] = "CANLibrary"
    custom_kw["D4"] = "Setter"
    custom_kw["E4"] = "Sends a signal value on the CAN bus"
    custom_kw["F4"] = "CANLibrary.Send('ACC_SetSpeed', 100)"
    custom_kw["C6"] = "ModelLibrary"
    custom_kw["D6"] = "Getter"
    custom_kw["E6"] = "Reads a model output value"
    custom_kw["F6"] = "ModelLibrary.GetOutput('ACC_ThrottleCmd')"

    wb.save(FIXTURES_DIR / "TE_TMHC_HILLS_Development & Testing_Keyword_Library_Description_Sheet.xlsx")


def main():
    build_system_requirements()
    build_command_list()
    build_configuration_file()
    build_compound_commands()
    build_library_description()
    print(f"Fixtures written to {FIXTURES_DIR}")


if __name__ == "__main__":
    main()
